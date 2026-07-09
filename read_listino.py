import sys
import json
try:
    import openpyxl
except ImportError:
    try:
        import pandas as pd
    except ImportError:
        print("openpyxl and pandas not found. Please install openpyxl.")
        sys.exit(1)

file_path = r"C:\Users\alber\Downloads\Listino_BMW_Noleggio_Premium.xlsx"

try:
    if 'openpyxl' in sys.modules:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            print(f"=== SHEET: {sheet_name} ===")
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                if any(row):
                    print([str(cell).strip() if cell is not None else "" for cell in row])
    else:
        df = pd.read_excel(file_path)
        print(df.to_string())
except Exception as e:
    print("Error reading file:", e)
